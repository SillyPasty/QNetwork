from top import Node, Edge
from openpyxl import load_workbook


def readExcelInput(path):
    """ 根据excel读取数据
    
    Args:
        path (str): 文件路径
    
    Returns
        list: 节点数组
    """
    wb = load_workbook(path)
    lengthSheet = wb.get_sheet_by_name('length')
    max_row = lengthSheet.max_row
    nodeList = []

    for i in range(0, max_row):
        node = Node(i)
        nodeList.append(node)

    for outNode in range(0, max_row):
        for inNode in range(0, max_row):
            cellValue = lengthSheet.cell(row=outNode+1, column=inNode+1).value
            if cellValue:
                nodeList[outNode].con(nodeList[inNode], cellValue)

    return nodeList


def recover(nodeList):
    """ 还原节点信息

    Args:
        nodeList (list): 节点列表
    
    """
    for node in nodeList:
        node.res = 0
        node.reling = 0
        node.path = []
        node.jumps = 0


def printTop(nodeList):
    """ 打印节点拓扑结构
    
    Args:
        nodeList (list): 节点列表
    
    """
    for node in nodeList:
        print('Node id %d ' % (node.nodeId + 1))
        adjNodesIdx = node.getAdjNodesIdx()
        for idx in adjNodesIdx:
            edges = node.getAdjEdge()[idx]
            print('node:%d length:%d' % ((idx + 1), edges.length()), end=',')
        print()


def initTop():
    nodeList = []
    size = int(input('Input the number of the nodes:'))

    for i in range(0, size):
        node = Node(i)
        nodeList.append(node)

    for outNode in range(0, size):
        row = input('Input the %d row:' % (outNode + 1))
        lens = row.split(' ')
        for inNode, length in enumerate(lens):
            length = int(length)
            if length and inNode != outNode:
                nodeList[outNode].con(nodeList[inNode], length)
    
    return nodeList